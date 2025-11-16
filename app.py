import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import io

# CSV読み込み
uploaded_file = "2025-11-15T04-52_export.csv"
df = pd.read_csv(uploaded_file)

# 列を文字列型に変換（「:」入力可能にする）
for col in ["勤務開始", "勤務終了", "休憩開始", "休憩終了"]:
    df[col] = df[col].astype(str)

st.set_page_config(page_title="昼休みシフト最適化ツール", layout="wide")
st.title("昼休みシフト最適化ツール")

st.subheader("スタッフ情報（CSVから読み込み済み）")
edited_data = st.data_editor(
    df,
    column_config={
        "休憩開始": st.column_config.TextColumn("休憩開始"),
        "休憩終了": st.column_config.TextColumn("休憩終了")
    },
    num_rows="dynamic"
)

break_duration_hours = st.number_input("休憩時間（時間）※未入力時のみ適用", min_value=1.0, max_value=3.0, value=2.0)

if st.button("スケジュールを作成"):
    staff_info = edited_data.copy()
    today = datetime.today().strftime("%Y-%m-%d")

    # 勤務時間をdatetimeに変換
    staff_info["勤務開始"] = pd.to_datetime(today + " " + staff_info["勤務開始"])
    staff_info["勤務終了"] = pd.to_datetime(today + " " + staff_info["勤務終了"])

    breaks = []
    for _, row in staff_info.iterrows():
        if str(row["休憩要否"]).lower() == "true":
            # 手動入力チェック（NaNや空欄を除外）
            if pd.notna(row["休憩開始"]) and pd.notna(row["休憩終了"]) and row["休憩開始"].strip() and row["休憩終了"].strip():
                try:
                    start = pd.to_datetime(today + " " + row["休憩開始"])
                    end = pd.to_datetime(today + " " + row["休憩終了"])
                    breaks.append((start, end))
                except ValueError:
                    breaks.append(None)
            else:
                # 自動割り当て（勤務時間が休憩時間より長い場合のみ）
                if (row["勤務終了"] - row["勤務開始"]).total_seconds()/3600 >= break_duration_hours:
                    start_range = pd.date_range(today + " 09:00", today + " 17:00", freq="30min")
                    start_range = [t for t in start_range if row["勤務開始"] <= t <= (row["勤務終了"] - pd.Timedelta(hours=break_duration_hours))]
                    if start_range:
                        best_start = start_range[int(len(start_range)/2)]
                        breaks.append((best_start, best_start + pd.Timedelta(hours=break_duration_hours)))
                    else:
                        breaks.append(None)
                else:
                    breaks.append(None)
        else:
            breaks.append(None)

    # タイムライン用データ
    timeline_data = []
    for i, row in enumerate(staff_info.itertuples()):
        # 勤務時間
        timeline_data.append({
            "スタッフ": row.スタッフ名,
            "開始": row.勤務開始,
            "終了": row.勤務終了,
            "タイプ": "勤務",
            "ラベル": f"{row.勤務開始.strftime('%H:%M')} - {row.勤務終了.strftime('%H:%M')}"
        })
        # 休憩時間
        if breaks[i] is not None:
            start, end = breaks[i]
            timeline_data.append({
                "スタッフ": row.スタッフ名,
                "開始": start,
                "終了": end,
                "タイプ": "休憩",
                "ラベル": f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"
            })

    timeline_df = pd.DataFrame(timeline_data)

    # タイムラインチャート
    fig = px.timeline(timeline_df, x_start="開始", x_end="終了", y="スタッフ", color="タイプ",
                      text="ラベル", color_discrete_map={"勤務": "seagreen", "休憩": "lightgray"})
    fig.update_traces(textposition="inside", insidetextanchor="middle", textfont=dict(size=14))
    fig.update_xaxes(range=[pd.to_datetime(today + " 08:00"), pd.to_datetime(today + " 18:00")], tickformat="%H:%M")
    fig.update_layout(height=600, title_text="勤務時間＋休憩時間（手動入力対応＋横軸固定）", font=dict(family="wqy-zenhei", size=16))

    st.plotly_chart(fig, use_container_width=True)

    # Excel出力
    excel_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "スケジュール"
    ws.append(["スタッフ", "休憩開始", "休憩終了"])
    for i, row in enumerate(staff_info.itertuples()):
        if breaks[i] is not None:
            ws.append([row.スタッフ名, breaks[i][0].strftime("%H:%M"), breaks[i][1].strftime("%H:%M")])
        else:
            ws.append([row.スタッフ名, "-", "-"])
    table = Table(displayName="Schedule", ref=f"A1:C{len(staff_info)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    for col in range(1, 4):
        max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    wb.save(excel_buffer)
    st.download_button("Excelファイルをダウンロード", data=excel_buffer.getvalue(), file_name="schedule.xlsx")