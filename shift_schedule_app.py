
import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import io

# ページ設定
st.set_page_config(page_title="昼休みシフト最適化ツール", layout="wide")
st.title("昼休みシフト最適化ツール")

# ✅ 初期データ（休憩要否は「要」→True、「否」→False）
data = {
    "スタッフ名": ["売店1", "売店2", "パート1", "パート2", "パート3", "パート4", "海運", "窓口1", "窓口2"],
    "勤務開始": ["08:00", "08:00", "09:00", "14:00", "09:00", "10:00", "08:00", "08:00", "08:00"],
    "勤務終了": ["18:00", "18:00", "12:00", "18:00", "15:00", "15:00", "18:00", "18:00", "18:00"],
    "休憩開始": ["", "", "", "", "", "", "", "", ""],
    "休憩終了": ["", "", "", "", "", "", "", "", ""],
    "休憩要否": [True, True, False, False, False, False, True, True, True]
}

df = pd.DataFrame(data)

# 列を文字列型に変換（勤務時間系）
for col in ["勤務開始", "勤務終了", "休憩開始", "休憩終了"]:
    df[col] = df[col].astype(str)

st.subheader("スタッフ情報（初期値固定）")

# 編集可能テーブル（休憩要否はチェックボックス）
edited_data = st.data_editor(
    df,
    column_config={
        "休憩開始": st.column_config.TextColumn("休憩開始"),
        "休憩終了": st.column_config.TextColumn("休憩終了"),
        "休憩要否": st.column_config.CheckboxColumn("休憩要否")
    },
    num_rows="dynamic"
)

# 休憩時間設定
break_duration_hours = st.number_input("休憩時間（時間）※未入力時のみ適用", min_value=1.0, max_value=3.0, value=2.0)

# スケジュール作成ボタン
if st.button("スケジュールを作成"):
    staff_info = edited_data.copy()
    today = datetime.today().strftime("%Y-%m-%d")

    # 勤務時間をdatetimeに変換
    staff_info["勤務開始"] = pd.to_datetime(today + " " + staff_info["勤務開始"])
    staff_info["勤務終了"] = pd.to_datetime(today + " " + staff_info["勤務終了"])

    breaks = []
    for _, row in staff_info.iterrows():
        if row["休憩要否"]:  # Trueなら休憩必要
            # 手動入力チェック
            if pd.notna(row["休憩開始"]) and pd.notna(row["休憩終了"]) and row["休憩開始"].strip() and row["休憩終了"].strip():
                try:
                    start = pd.to_datetime(today + " " + row["休憩開始"])
                    end = pd.to_datetime(today + " " + row["休憩終了"])
                    breaks.append((start, end))
                except ValueError:
                    breaks.append(None)
            else:
                # ✅ 改良版：昼休みを12:00中心に割り当て
                if (row["勤務終了"] - row["勤務開始"]).total_seconds() / 3600 >= break_duration_hours:
                    preferred_start = pd.to_datetime(today + " 12:00")
                    start = max(row["勤務開始"], preferred_start)
                    end = start + pd.Timedelta(hours=break_duration_hours)
                    if end <= row["勤務終了"]:
                        breaks.append((start, end))
                    else:
                        # fallback: 勤務時間の中間
                        mid_start = row["勤務開始"] + (row["勤務終了"] - row["勤務開始"]) / 2 - pd.Timedelta(hours=break_duration_hours / 2)
                        breaks.append((mid_start, mid_start + pd.Timedelta(hours=break_duration_hours)))
                else:
                    breaks.append(None)
        else:
            breaks.append(None)

    # タイムライン用データ
    timeline_data = []
    for i, row in enumerate(staff_info.itertuples()):
        timeline_data.append({
            "スタッフ": row.スタッフ名,
            "開始": row.勤務開始,
            "終了": row.勤務終了,
            "タイプ": "勤務",
            "ラベル": f"{row.勤務開始.strftime('%H:%M')} - {row.勤務終了.strftime('%H:%M')}"
        })
        if breaks[i] is not None:
            start, end = breaks[i]
            timeline_data.append({
                "スタッフ": row.スタッフ名,
                "開始": start,
                "終了": end,
                "タイプ": "休憩",
                "ラベル": f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}"
            })

    timeline_df = pd.DataFrame(timeline_data)

    # タイムラインチャート
    fig = px.timeline(
        timeline_df, x_start="開始", x_end="終了", y="スタッフ", color="タイプ",
        text="ラベル", color_discrete_map={"勤務": "seagreen", "休憩": "lightgray"}
    )
    fig.update_traces(textposition="inside", insidetextanchor="middle", textfont=dict(size=14))
    fig.update_xaxes(range=[pd.to_datetime(today + " 08:00"), pd.to_datetime(today + " 18:00")], tickformat="%H:%M")
    fig.update_layout(height=600, title_text="勤務時間＋休憩時間（昼休み優先）", font=dict(family="wqy-zenhei", size=16))
    st.plotly_chart(fig, use_container_width=True)

    # Excel出力
    excel_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "スケジュール"
    ws.append(["スタッフ", "休憩開始", "休憩終了"])
    for i, row in enumerate(staff_info.itertuples()):
        if breaks[i] is not None:
            ws.append([row.スタッフ名, breaks[i][0].strftime("%H:%M"), breaks[i][1].strftime("%H:%M")])
        else:
            ws.append([row.スタッフ名, "-", "-"])
    table = Table(displayName="Schedule", ref=f"A1:C{len(staff_info)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    for col in range(1, 4):
        max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2
    wb.save(excel_buffer)
    st.download_button("Excelファイルをダウンロード", data=excel_buffer.getvalue(), file_name="schedule.xlsx")
